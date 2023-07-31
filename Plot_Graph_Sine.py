import numpy as np
from PyQt5.QtWidgets import *
from PyQt5 import uic
import sys

import time


import pickle
import math
import openpyxl
import os

import pandas as pd
from matplotlib import pyplot as plt  # import matplotlib.pyplot as plt
from pyqtgraph import PlotWidget

from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.backends.backend_qt5agg import NavigationToolbar2QT as NavigationToolbar
from matplotlib.figure import Figure
from PyQt5.QtWidgets import QCheckBox


from PyQt5.QtWidgets import QDialog, QApplication, QPushButton, QVBoxLayout, QShortcut

from matplotlib.animation import FuncAnimation

import matplotlib
matplotlib.use('TkAgg')

from scipy.interpolate import CubicSpline

import keyboard

from PyQt5.QtGui import QKeySequence

from PyQt5.QtWidgets import QApplication, QMainWindow, QVBoxLayout, QWidget, QPushButton, QFileDialog, QLabel

class sine_graph():
    
    def __init__(self):
        
        self.figure = plt.figure()              #       self.figure = Figure()          both are same may be

        plt.tight_layout()
        self.canvas = FigureCanvas(self.figure)
        self.toolbar = NavigationToolbar(self.canvas, None)

        self.graph_axes = self.figure.add_subplot(111)
        
        self.displacement_list = []
        self.Points_list = []
        self.time_list = []
        # self.graph_axes.cla()
        pass
    
    def plot_now(self,Points, CC_Text, Duty_Text, Beat_Rate_Text):
    
        self.Points = Points
        self.CC_Text = CC_Text
        self.Duty_Text = Duty_Text
        self.Beat_Rate_Text = Beat_Rate_Text

        time_per_beat = 60 / self.Beat_Rate_Text
        print("time_per_beat",time_per_beat)
        
        time_per_point = time_per_beat / Points
        print("time_per_point",time_per_point)

        print("i am here at 229")

        print("time_per_beat", time_per_beat)
        # print("time_per_point",time_per_point)						#	temporary comment

        Duty_cycle_digit = 100 / self.Duty_Text
        print("Duty_cycle_digit",Duty_cycle_digit)
        
        forwad_cycle = int(Points / Duty_cycle_digit)
        Reverse_cycle = Points - forwad_cycle

        on_time = forwad_cycle * time_per_point
        off_time = Reverse_cycle * time_per_point

        print("")
        # print("Duty_cycle_digit",Duty_cycle_digit)						#	temporary comment
        print("forwad_cycle", forwad_cycle)
        print("Reverse_cycle", Reverse_cycle)

        print("")
        print("on_time", on_time)
        print("off_time", off_time)

        forwad_cycle_points_with_degree = 180 / forwad_cycle
        Reverse_cycle_points_with_degree = 180 / Reverse_cycle

        degree = 90
        degree_list = []
        radian_list = []

        Points_initial = 0

        self.displacement_list = []               #       self.displacement_list = ["Displacement"]
        self.Points_list = []                           #       self.Points_list = ["Points"]
        self.time_list = []                          #       self.time_list = ["Time (ms)"] 

        for i in range(forwad_cycle):
            degree = degree + forwad_cycle_points_with_degree
            degree_list.append(degree)

            radian = degree*(3.141592/180)
            radian_list.append(radian)

            displacement = round(self.CC_Text * math.sin(radian))
            self.displacement_list.append(displacement)

            Points_initial = Points_initial + 1
            self.Points_list.append(Points_initial)
            self.time_list.append((round((time_per_point*Points_initial*1000), 2)))

        for i in range(Reverse_cycle):
            degree = degree + Reverse_cycle_points_with_degree
            degree_list.append(degree)

            radian = degree * (3.141592/180)
            radian_list.append(radian)

            displacement = round(self.CC_Text * math.sin(radian))
            self.displacement_list.append(displacement)

            Points_initial = Points_initial + 1
            self.Points_list.append(Points_initial)
            self.time_list.append((round((time_per_point*Points_initial*1000), 2)))


        self.x = self.Points_list
        self.y = self.displacement_list
        self.z = self.time_list
        
        ##################################################################################################################################################

        # temp = [self.x, self.y, self.z]
        
        
        # with open("Final_Code_P", 'wb') as outfile:
        #     pickle.dump(temp, outfile, pickle.HIGHEST_PROTOCOL)

        # with open('Final_Code_P', 'rb') as file:
        #     data = pickle.load(file)
            
        # ##################################################################################################################################################
        
        # workbook = openpyxl.Workbook()

        # sheet = workbook.active

        # # Write the data to the Excel file
        # for row_index, row_data in enumerate(data, start=1):
        #     for col_index, cell_data in enumerate(row_data, start=1):
        #         sheet.cell(row=col_index, column=row_index, value=cell_data)

        # # Save the workbook
        # workbook.save('Final_Code_P.xlsx')

        # # os.startfile("C:\\Users\\divyesh.hadvani\\Desktop\\python_code_AH\\Trial_01\\Final_Code_P.xlsx")
        
        ##################################################################################################################################################

        self.plot_graph(self.y,self.z)
        
    def plot_graph(self,y,z):
        
        x_values = z
        y_values = y

        self.graph_axes.clear()

        # Plot the data
        self.graph_axes.plot(x_values, y_values)
        self.graph_axes.grid(True)
        self.graph_axes.set_xlabel('Time in milli sec')
        self.graph_axes.set_ylabel('Distance in micro meter')
        self.graph_axes.set_title('Distance vs Time')
        plt.tight_layout()

        # Update the canvas
        self.figure.canvas.draw()
        # self.figure.canvas.flush_events()
        
    def data_export(self,Beat_Rate=0, CC=0, Duty_Cycle=0, Points_list=[], displacement_list=[], time_list=[]):
        
        self.Beat_Rate = Beat_Rate
        self.CC = CC
        self.Duty_Cycle = Duty_Cycle     
        self.Points_list = Points_list
        self.displacement_list = displacement_list
        self.time_list = time_list
        
        ##################################################################################################################################################
        
        Input_Name_List = ["Beat Rate", "CC", "Duty Cycle"]
        Input_Data_List = [self.Beat_Rate, self.CC, self.Duty_Cycle]

        a = Input_Name_List
        b = Input_Data_List
        # displacement_list = ["Displacement"]
        # Points_list = ["Points"]
        # time_list = ["Time (ms)"]

        x = Points_list
        y = displacement_list
        z = time_list
        
        
        
        # Find the length of the longest array
        max_length = max(len(a), len(b), len(x), len(y), len(z))

        # Pad the shorter arrays with NaN to match the length of the longest array
        a += [float('nan')] * (max_length - len(a))
        b += [float('nan')] * (max_length - len(b))
        x += [float('nan')] * (max_length - len(x))
        y += [float('nan')] * (max_length - len(y))
        z += [float('nan')] * (max_length - len(z))
        
        data = {
            'Parameters Name': a,
            'Parameters Value': b,
            'Points': x,
            'Displacement': y,
            'Time (ms)': z
            }
        
        df = pd.DataFrame(data)
        
        # Get the file path using a save dialog box
        file_path, _ = QFileDialog.getSaveFileName(
            caption='Save Excel File', directory='data.xlsx', filter='Excel Files (*.xlsx)')
        
        # Check if the user provided a file path
        if file_path:
            # Save the DataFrame to the Excel file
            df.to_excel(file_path, index=False)
            print(f"Data saved to {file_path}")
        else:
            print("Save canceled")
            
        os.startfile(file_path)      
        
        # self.save_file()
        
        
        ##################################################################################################################################################
        
        # Input_Name_List = ["Parameters Name","Beat Rate", "CC", "Duty Cycle"]
        # Input_Data_List = ["Parameters Value",self.Beat_Rate, self.CC, self.Duty_Cycle]

        # a = Input_Name_List
        # b = Input_Data_List
        # # displacement_list = ["Displacement"]
        # # Points_list = ["Points"]
        # # time_list = ["Time (ms)"]

        # x = Points_list
        # y = displacement_list
        # z = time_list

        # temp = [a, b, x, y, z]

        # with open("Export_Sine_File", 'wb') as outfile:
        #     pickle.dump(temp, outfile, pickle.HIGHEST_PROTOCOL)

        # with open('Export_Sine_File', 'rb') as file:
        #     data = pickle.load(file)

        # workbook = openpyxl.Workbook()

        # sheet = workbook.active
        
        # # print("data")
        # # print(data)
        # # print("#####################################################################")

        # # Write the data to the Excel file
        # for row_index, row_data in enumerate(data, start=1):
        #     # print("row_index")
        #     # print(row_index)
        #     # print("#####################################################################")
        #     # print("row_data")
        #     # print(row_data)
        #     # print("#####################################################################")
            
        #     for col_index, cell_data in enumerate(row_data, start=1):
        #         sheet.cell(row=col_index, column=row_index,
        #                    value=cell_data)

        # # Save the workbook
        # workbook.save('Export_Sine_File.xlsx')
        
        # os.startfile(
        #     "C:\\Users\\divyesh.hadvani\\Desktop\\python_code_AH\\Trial_01\\Export_Sine_File.xlsx")  
        
        ##################################################################################################################################################

        
        
    ####################################################################################################################################################################################################
    ####################################################################################################################################################################################################
    ####################################################################################################################################################################################################
    ####################################################################################################################################################################################################
        
        
    def generate_points_by_click(self, Points, CC_Text, Duty_Text, Beat_Rate_Text):
        
        self.export_pressed = False
        
        self.graph_axes.clear()
        
        self.Points = Points
        self.CC_Text = CC_Text
        self.Duty_Text = Duty_Text
        self.Beat_Rate_Text = Beat_Rate_Text
        
        time_per_beat = 60 / self.Beat_Rate_Text
        print("time_per_beat",time_per_beat)
        
        time_per_point = time_per_beat / Points
        print("time_per_point",time_per_point)

        print("i am here at 229")

        print("time_per_beat", time_per_beat)
        # print("time_per_point",time_per_point)						#	temporary comment

        Duty_cycle_digit = 100 / self.Duty_Text
        print("Duty_cycle_digit",Duty_cycle_digit)
        
        forwad_cycle = int(Points / Duty_cycle_digit)
        Reverse_cycle = Points - forwad_cycle

        on_time = forwad_cycle * time_per_point
        off_time = Reverse_cycle * time_per_point

        print("")
        # print("Duty_cycle_digit",Duty_cycle_digit)						#	temporary comment
        print("forwad_cycle", forwad_cycle)
        print("Reverse_cycle", Reverse_cycle)

        print("")
        print("on_time", on_time)
        print("off_time", off_time)
        
        self.x_n_value = time_per_beat * 1000        
        
        # time_per_beat = 60 / self.Beat_Rate_Text
        # time_per_point = time_per_beat / Points

        displacement_list = []
        Points_list = []
        main_list = [(0, self.CC_Text), ((self.x_n_value/Duty_cycle_digit), 0),
                     (self.x_n_value, self.CC_Text)]
        # main_list = [(0,10),(5,0),(10,10)]

        self.graph_axes.set_xlim(0, self.x_n_value)
        self.graph_axes.set_ylim(0, self.CC_Text)

        # Plot an empty scatter plot
        scatter = self.graph_axes.scatter([], [])

        while True:
                        
            # if self.Export_PB.clicked == True:
            if self.export_pressed == True:
                print("i am herer")
                break

            # Get mouse click input
            point = plt.ginput(n=1, timeout=0.5)
            try:
                if not point:
                    # No click occurred
                    continue

                if point[0][0] < 0 or point[0][1] < 0:
                    # Negative coordinates, ignore
                    continue

                if point[0][0] > self.x_n_value or point[0][1] > self.CC_Text:
                    # Coordinates out of range, ignore
                    continue

                # Add the clicked point to the scatter plot
                scatter.set_offsets(point)

                # Print the coordinates of the clicked point
                print("Point:", point[0])
                main_list.append(point[0])

                try:
                    # print("main_list",main_list)
                    sorted_list = sorted(main_list, key=lambda x: x[0])
                    # print("sorted_list",sorted_list)
                    random_points = sorted_list

                    for i in sorted_list:
                        Points_list.append(i[0])
                        displacement_list.append(i[1])

                    # plt.scatter(Points_list, displacement_list, color='blue')
                    
                    self.graph_axes.scatter(Points_list, displacement_list, color='blue')

                    # Separate x and y coordinates from the given points
                    x_coordinates, y_coordinates = zip(*random_points)
                    # print(x_coordinates)

                    # Create a cubic spline interpolator
                    spline = CubicSpline(x_coordinates, y_coordinates)

                    # Generate x values for the plot
                    x_plot = np.linspace(min(x_coordinates),
                                        max(x_coordinates), 100)
                    y_plot = spline(x_plot)

                    # plt.clf()
                    # plt.xlim(0, self.Beat_Rate_Text)
                    # plt.ylim(0, self.CC_Text)
                    
                    self.graph_axes.clear()
                    self.graph_axes.set_xlim(0, self.x_n_value)
                    self.graph_axes.set_ylim(0, self.CC_Text)

                    # Plot the data points and the cubic polynomial curve
                    # plt.scatter(x_coordinates, y_coordinates,
                    #             color='blue', label='Data Points')
                    # plt.plot(x_plot, y_plot, color='red', label='Cubic Spline')
                    
                    self.graph_axes.scatter(x_coordinates, y_coordinates,
                                color='blue', label='Data Points')
                    self.graph_axes.plot(x_plot, y_plot, color='red', label='Cubic Spline')

                    self.graph_axes.grid(True)
                    self.graph_axes.set_xlabel('Time in milli sec')
                    self.graph_axes.set_ylabel('Distance in micro meter')
                    self.graph_axes.set_title('Distance vs Time')
                    plt.tight_layout()                                                  #   self.graph_axes.tight_layout()      is not working

                    # plt.xlabel('Points_Time')  # added by dev
                    # plt.ylabel('Displacement')  # added by dev
                    # plt.grid(True)  # added by dev
                    # plt.title('Cubic Spline Graph')

                    # Add legend and grid
                    # plt.legend()

                except:
                    print("waiting for 2nd point")
  
            except:
                pass
            
        ##################################################################################################################################################
        
        Input_Name_List = ["Beat Rate", "CC", "Duty Cycle"]
        Input_Data_List = [self.Beat_Rate_Text, self.CC_Text, self.Duty_Text]

        a = Input_Name_List
        b = Input_Data_List
        x = Points_list
        y = y_plot.tolist()
        z = x_plot.tolist()
        
        
        
        # Find the length of the longest array
        max_length = max(len(a), len(b), len(x), len(y), len(z))

        # Pad the shorter arrays with NaN to match the length of the longest array
        a += [float('nan')] * (max_length - len(a))
        b += [float('nan')] * (max_length - len(b))
        x += [float('nan')] * (max_length - len(x))
        y += [float('nan')] * (max_length - len(y))
        z += [float('nan')] * (max_length - len(z))
        
        data = {
            'Parameters Name': a,
            'Parameters Value': b,
            'Points at Time (ms)': x,
            'Displacement': y,
            'Time (ms)': z
            }
        
        df = pd.DataFrame(data)
        
        # Get the file path using a save dialog box
        file_path, _ = QFileDialog.getSaveFileName(
            caption='Save Excel File', directory='data.xlsx', filter='Excel Files (*.xlsx)')
        
        # Check if the user provided a file path
        if file_path:
            # Save the DataFrame to the Excel file
            df.to_excel(file_path, index=False)
            print(f"Data saved to {file_path}")
        else:
            print("Save canceled")
            
        os.startfile(file_path)      
        
        # self.save_file()
        
        
        ##################################################################################################################################################

        # # plt.close()

        # workbook = openpyxl.Workbook()

        # sheet = workbook.active
        
        # Input_Name_List = ["Beat Rate", "CC", "Duty Cycle"]
        # Input_Data_List = [self.Beat_Rate_Text, self.CC_Text, self.Duty_Text]

        # # displacement_list = ["Displacement"]
        # # Points_list = ["Points"]
        # # time_list = ["Time (ms)"]

        # print("i am here at 433")

        # list01 = x_plot.tolist()
        # list02 = y_plot.tolist()

        # list1 = ["Points_Time"] + list01
        # list2 = ["displacement"] + list02

        # # Write data to the Excel file
        # for index, value in enumerate(Input_Name_List, start=1):
        #     sheet.cell(row=index, column=1, value=value)
            
        # for index, value in enumerate(Input_Data_List, start=1):
        #     sheet.cell(row=index, column=2, value=value)
            
        # for index, value in enumerate(list2, start=1):
        #     sheet.cell(row=index, column=3, value=value)
            
        # for index, value in enumerate(list1, start=1):
        #     sheet.cell(row=index, column=4, value=value)

        # # Save the workbook
        # workbook.save('cubic_spline_with_random_points.xlsx')

        # os.startfile(
        #     "C:\\Users\\divyesh.hadvani\\Desktop\\python_code_AH\\Trial_01\\cubic_spline_with_random_points.xlsx")
        
        
        ##################################################################################################################################################
        
    def Export_PB_CLICKED(self):
        print("i am in export")
        self.export_pressed = True
        
    
    ####################################################################################################################################################################################################
